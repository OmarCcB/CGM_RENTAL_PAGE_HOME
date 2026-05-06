import sqlite3, os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

conn = sqlite3.connect('cgm.db')
conn.row_factory = sqlite3.Row
cur = conn.cursor()
cur.execute('SELECT nombre, tipo, slug, imagen, ficha_url, activo, show_arg FROM products ORDER BY show_arg, activo DESC, tipo, nombre')
rows = cur.fetchall()

def img_status(img):
    if not img or not img.strip(): return ('NO', 'Campo vacio')
    full = os.path.join('static', 'products', img)
    if os.path.exists(full):
        files = [f for f in os.listdir(os.path.dirname(full)) if not f.startswith('.') and f != '.gitkeep']
        return ('SI', 'OK') if files else ('NO', 'Carpeta vacia')
    return ('NO', 'Archivo no existe')

def ficha_status(ficha):
    if not ficha or not ficha.strip(): return ('NO', 'Sin ficha')
    if ficha.startswith('http'): return ('EXT', 'Externa URL')
    full = os.path.join('static', 'docs', ficha)
    return ('SI', 'OK') if os.path.exists(full) else ('NO', 'Archivo no existe')

# Orphaned images
used_folders = set()
for r in rows:
    if r['imagen']: used_folders.add(r['imagen'].split('/')[0])
all_folders = set(os.listdir('static/products'))
orphan_folders = sorted(all_folders - used_folders)

# Orphaned fichas
used_fichas = set()
for r in rows:
    if r['ficha_url'] and not r['ficha_url'].startswith('http'):
        used_fichas.add(os.path.basename(r['ficha_url']))
all_fichas = set(os.listdir('static/docs/fichas'))
orphan_fichas = sorted(all_fichas - used_fichas)

VERDE_OSC = "1B4332"
VERDE_MED = "2D6A4F"
GRIS      = "6B7280"
ROJO_BG   = "FFD6D6"
ROJO_F    = "C0392B"
VERDE_BG  = "E8F5E9"
VERDE_F   = "1B5E20"
BLANCO    = "FFFFFF"

def fill(h): return PatternFill("solid", fgColor=h)
def brd():
    t = Side(style='thin', color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

wb = Workbook()

# ══ SHEET 1: Inventario ══
ws = wb.active
ws.title = "Inventario"

ws.merge_cells('A1:F1')
c = ws['A1']
c.value = "INVENTARIO EQUIPOS CGM RENTAL"
c.font = Font(name='Calibri', bold=True, size=13, color=BLANCO)
c.fill = fill(VERDE_OSC)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 28

for col, h in enumerate(['EQUIPO','CATEGORIA','PAIS','ESTADO','IMAGEN','FICHA TECNICA'], 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = Font(name='Calibri', bold=True, size=10, color=BLANCO)
    cell.fill = fill(VERDE_MED)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = brd()
ws.row_dimensions[2].height = 20

group_bg = {('PE',True):"F0FFF4",('PE',False):"FFF8F0",('AR',True):"F0F8FF",('AR',False):"FFF0F0"}
current_group = None
row_num = 3

for r in rows:
    pais   = 'AR' if r['show_arg']==1 else 'PE'
    activo = r['activo'] == 1
    group  = (pais, activo)

    if group != current_group:
        current_group = group
        lbl = ("PERU" if pais=='PE' else "ARGENTINA") + "  -  " + ("ACTIVOS" if activo else "INACTIVOS")
        ws.merge_cells(f'A{row_num}:F{row_num}')
        sec = ws.cell(row=row_num, column=1, value=lbl)
        sec.font = Font(name='Calibri', bold=True, size=10, color=BLANCO)
        sec.fill = fill(VERDE_OSC if activo else GRIS)
        sec.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row_num].height = 16
        row_num += 1

    bg = group_bg.get(group, BLANCO)
    i_ok, i_txt = img_status(r['imagen'])
    f_ok, f_txt = ficha_status(r['ficha_url'])

    data = [r['nombre'], r['tipo'], pais, 'Activo' if activo else 'Inactivo', i_txt, f_txt]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = Font(name='Calibri', size=9)
        cell.border = brd()
        cell.alignment = Alignment(vertical='center', wrap_text=(col==1))
        cell.fill = fill(bg)
        if col == 5:
            cell.fill = fill(VERDE_BG if i_ok=='SI' else ROJO_BG)
            cell.font = Font(name='Calibri', size=9, color=VERDE_F if i_ok=='SI' else ROJO_F, bold=(i_ok!='SI'))
        if col == 6:
            if f_ok == 'SI':
                cell.fill = fill(VERDE_BG); cell.font = Font(name='Calibri', size=9, color=VERDE_F)
            elif f_ok == 'NO':
                cell.fill = fill(ROJO_BG); cell.font = Font(name='Calibri', size=9, color=ROJO_F, bold=True)
        if col == 3:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=9, bold=True, color=VERDE_OSC if pais=='PE' else "003399")
        if col == 4:
            cell.font = Font(name='Calibri', size=9, color=VERDE_F if activo else GRIS, bold=activo)
    ws.row_dimensions[row_num].height = 15
    row_num += 1

ws.column_dimensions['A'].width = 44
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 18
ws.freeze_panes = 'A3'

# ══ SHEET 2: Archivos sin uso ══
ws2 = wb.create_sheet("Archivos Sin Uso")

ws2.merge_cells('A1:C1')
c = ws2['A1']
c.value = "ARCHIVOS SIN USO — no referenciados por ningun equipo en la DB"
c.font = Font(name='Calibri', bold=True, size=11, color=BLANCO)
c.fill = fill(VERDE_OSC)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 22

row2 = 2
for hdr in [("CARPETAS DE IMAGENES SIN USO", "N. ARCHIVOS")]:
    ws2.cell(row=row2, column=1, value=hdr[0]).font = Font(bold=True, size=10, color=BLANCO, name='Calibri')
    ws2.cell(row=row2, column=1).fill = fill(GRIS)
    ws2.cell(row=row2, column=2, value=hdr[1]).font = Font(bold=True, size=10, color=BLANCO, name='Calibri')
    ws2.cell(row=row2, column=2).fill = fill(GRIS)
    row2 += 1

for f in orphan_folders:
    path = os.path.join('static', 'products', f)
    files = [x for x in os.listdir(path) if not x.startswith('.') and x != '.gitkeep']
    ws2.cell(row=row2, column=1, value=f).font = Font(name='Calibri', size=9)
    ws2.cell(row=row2, column=1).border = brd()
    ws2.cell(row=row2, column=2, value=len(files)).font = Font(name='Calibri', size=9)
    ws2.cell(row=row2, column=2).border = brd()
    ws2.cell(row=row2, column=2).alignment = Alignment(horizontal='center')
    row2 += 1

row2 += 1
ws2.cell(row=row2, column=1, value="FICHAS PDF SIN USO").font = Font(bold=True, size=10, color=BLANCO, name='Calibri')
ws2.cell(row=row2, column=1).fill = fill(GRIS)
ws2.cell(row=row2, column=2, value="TAMANO (KB)").font = Font(bold=True, size=10, color=BLANCO, name='Calibri')
ws2.cell(row=row2, column=2).fill = fill(GRIS)
row2 += 1

for f in orphan_fichas:
    size = os.path.getsize(os.path.join('static','docs','fichas',f)) // 1024
    ws2.cell(row=row2, column=1, value=f).font = Font(name='Calibri', size=9)
    ws2.cell(row=row2, column=1).border = brd()
    ws2.cell(row=row2, column=2, value=size).font = Font(name='Calibri', size=9)
    ws2.cell(row=row2, column=2).border = brd()
    ws2.cell(row=row2, column=2).alignment = Alignment(horizontal='center')
    row2 += 1

ws2.column_dimensions['A'].width = 55
ws2.column_dimensions['B'].width = 15

wb.save('inventario_equipos_cgm.xlsx')
print(f"Guardado OK")
print(f"Carpetas de imagenes sin uso: {len(orphan_folders)}")
for f in orphan_folders:
    path = os.path.join('static','products',f)
    n = len([x for x in os.listdir(path) if not x.startswith('.') and x != '.gitkeep'])
    print(f"  {f} ({n} archivos)")
print(f"Fichas PDF sin uso: {len(orphan_fichas)}")
for f in orphan_fichas:
    print(f"  {f}")
conn.close()
