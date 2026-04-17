import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# DATA: brochure desglosado — 1 fila por modelo
# Campos: (sector, tipo, equipo, marca, modelo, en_web, pais_web)
# en_web: 'SI-PER' | 'SI-ARG' | 'SI-AMBOS' | 'NO'
# ══════════════════════════════════════════════════════════════════════════════

data = [
    # ── CONSTRUCCIÓN ──────────────────────────────────────────────────────────
    ('Construcción', 'Excavadora',          'Excavadora Hidráulica',      'John Deere', '210G-LC',        'SI-AMBOS'),
    ('Construcción', 'Excavadora',          'Excavadora Hidráulica',      'Hitachi',    'ZX210',          'SI-PER'),
    ('Construcción', 'Excavadora',          'Excavadora Hidráulica',      'John Deere', '350G-LC',        'SI-AMBOS'),
    ('Construcción', 'Excavadora',          'Excavadora Hidráulica',      'John Deere', '350P-LC',        'SI-AMBOS'),
    ('Construcción', 'Excavadora',          'Excavadora Hidráulica',      'Hitachi',    'ZX350',          'SI-PER'),
    ('Construcción', 'Tractor de Orugas',   'Tractor de Orugas',          'John Deere', '850J-II',        'SI-AMBOS'),
    ('Construcción', 'Motoniveladora',      'Motoniveladora',             'John Deere', '620G',           'SI-AMBOS'),
    ('Construcción', 'Motoniveladora',      'Motoniveladora',             'John Deere', '670G',           'SI-PER'),
    ('Construcción', 'Motoniveladora',      'Motoniveladora',             'John Deere', '620P',           'SI-AMBOS'),
    ('Construcción', 'Cargador Frontal',    'Cargador Frontal',           'John Deere', '644K',           'SI-AMBOS'),
    ('Construcción', 'Cargador Frontal',    'Cargador Frontal',           'John Deere', '744K-II',        'SI-PER'),
    ('Construcción', 'Cargador Frontal',    'Cargador Frontal',           'John Deere', '744L',           'SI-PER'),
    ('Construcción', 'Cargador Frontal',    'Cargador Frontal',           'John Deere', '744P',           'SI-PER'),
    ('Construcción', 'Cargador Frontal',    'Cargador Frontal',           'John Deere', '844L',           'NO'),
    ('Construcción', 'Minicargador',        'Minicargador',               'John Deere', '324G',           'SI-AMBOS'),
    ('Construcción', 'Minicargador',        'Minicargador',               'John Deere', '320G',           'SI-AMBOS'),
    ('Construcción', 'Retroexcavadora',     'Retroexcavadora',            'John Deere', '310SL',          'SI-AMBOS'),
    ('Construcción', 'Retroexcavadora',     'Retroexcavadora',            'John Deere', '310L',           'SI-AMBOS'),
    ('Construcción', 'Retroexcavadora',     'Retroexcavadora',            'John Deere', '310P',           'SI-AMBOS'),
    ('Construcción', 'Retroexcavadora',     'Retroexcavadora',            'John Deere', '320P',           'SI-AMBOS'),
    ('Construcción', 'Rodillo Compactador', 'Rodillo Compactador',        'HAMM',       '3520 HT',        'SI-PER'),
    ('Construcción', 'Rodillo Compactador', 'Rodillo Compactador',        'HAMM',       'HC 200',         'SI-PER'),
    ('Construcción', 'Rodillo Compactador', 'Rodillo Compactador',        'HAMM',       '3411',           'SI-PER'),
    ('Construcción', 'Rodillo Compactador', 'Rodillo Compactador HC 100', 'HAMM',       'HC 100i C',      'SI-AMBOS'),
    ('Construcción', 'Rodillo Tándem',      'Rodillo Tándem',             'HAMM',       'HD 12 VV',       'SI-PER'),
    ('Construcción', 'Micropavimentadora',  'Micropavimentadora',         'Bergkamp',   'CF 85.460',      'NO'),
    ('Construcción', 'Camión Cisterna',     'Camión Cisterna',            'DAF',        'CF 410',         'SI-AMBOS'),
    ('Construcción', 'Camión Grúa',         'Camión Grúa',                'DAF',        'CF 410',         'SI-PER'),
    ('Construcción', 'Camión Volquete',     'Camión Volquete',            'DAF',        'CF 480',         'NO'),
    ('Construcción', 'Pavimentadora',       'Pavimentadora',              'Vogele',     'Super 1800-3i',  'NO'),
    ('Construcción', 'Pavimentadora',       'Pavimentadora',              'Vogele',     '1400',           'NO'),
    ('Construcción', 'Autohormigonera',     'Autohormigonera',            'Fiori',      'DB 460',         'NO'),
    ('Construcción', 'Chancadora',          'Chancadora Primaria',        'Kleemann',   'MOBICAT MC 110 EVO2', 'NO'),
    ('Construcción', 'Chancadora',          'Chancadora Secundaria',      'Kleemann',   'MOBICAT MC 90 EVO2',  'NO'),
    ('Construcción', 'Zaranda',             'Zaranda',                    'Kleemann',   '—',              'NO'),
    ('Construcción', 'Faja Transportadora', 'Faja Transportadora',        'Kleemann',   'MOBICAT MBT 24', 'NO'),

    # ── MINERÍA ───────────────────────────────────────────────────────────────
    ('Minería',      'Excavadora',          'Excavadora Hidráulica',      'John Deere', '870G',           'SI-PER'),
    ('Minería',      'Excavadora',          'Excavadora Hidráulica',      'Hitachi',    'EX1200',         'SI-PER'),

    # ── ENERGÍA ───────────────────────────────────────────────────────────────
    ('Energía',      'Generador',           'Generador',                  'AKSA',       'AK55CUST',       'SI-PER'),
    ('Energía',      'Generador',           'Generador',                  'AKSA',       'AK70CUST',       'NO'),
    ('Energía',      'Generador',           'Generador',                  'AKSA',       'AK90CUST',       'SI-PER'),
    ('Energía',      'Generador',           'Generador',                  'AKSA',       'AK160CUST',      'SI-PER'),
    ('Energía',      'Generador',           'Generador',                  'AKSA',       'AK320CUST',      'NO'),
    ('Energía',      'Generador',           'Generador',                  'AKSA',       'AK360CUST',      'SI-PER'),
    ('Energía',      'Generador',           'Generador',                  'AKSA',       'AK450CUST',      'SI-PER'),
    ('Energía',      'Generador',           'Generador',                  'AKSA',       'AK505VOLST',     'SI-PER'),
    ('Energía',      'Generador',           'Generador',                  'TEKSAN',     'TK55CUST',       'NO'),
    ('Energía',      'Generador',           'Generador',                  'TEKSAN',     'TK90PELS',       'SI-PER'),
    ('Energía',      'Generador',           'Generador',                  'TEKSAN',     'TK120PELS',      'SI-PER'),
    ('Energía',      'Generador',           'Generador',                  'TEKSAN',     'TK180PELS',      'NO'),
    ('Energía',      'Generador',           'Generador',                  'TEKSAN',     'TK220PELS',      'NO'),
    ('Energía',      'Generador',           'Generador',                  'TEKSAN',     'TK280PELS',      'NO'),
    ('Energía',      'Generador',           'Generador',                  'TEKSAN',     'TK360PELS',      'SI-PER'),
    ('Energía',      'Generador',           'Generador',                  'TEKSAN',     'TK470PELS',      'SI-PER'),
    ('Energía',      'Generador',           'Generador',                  'TEKSAN',     'TK500PELS',      'NO'),
    ('Energía',      'Generador',           'Generador',                  'TEKSAN',     'TK550PELS',      'SI-PER'),
    ('Energía',      'Generador',           'Generador',                  'TEKSAN',     'TK725CUST',      'NO'),
    ('Energía',      'Torre de Iluminación','Torre de Iluminación',       'Atlas Copco','Hilight V5+',    'SI-AMBOS'),
    ('Energía',      'Torre de Iluminación','Torre de Iluminación',       'Generac Vtevo','GLT4-M',       'NO'),
    ('Energía',      'Compresora',          'Compresora',                 'Atlas Copco','XAS186',         'NO'),

    # ── AGRÍCOLA ──────────────────────────────────────────────────────────────
    ('Agrícola',     'Tractor Grande',      'Tractor Agrícola',           'John Deere', '7M-230',         'NO'),
    ('Agrícola',     'Tractor Utilitarios', 'Tractor Agrícola',           'John Deere', '5090E',          'SI-PER'),
    ('Agrícola',     'Tractor Utilitarios', 'Tractor Agrícola',           'John Deere', '5090EN',         'SI-PER'),
    ('Agrícola',     'Tractor Utilitarios', 'Tractor Agrícola',           'John Deere', '5090EH',         'NO'),
    ('Agrícola',     'Tractor Utilitarios', 'Tractor Agrícola',           'John Deere', '5080EN',         'NO'),
    ('Agrícola',     'Tractor Utilitarios', 'Tractor Agrícola',           'John Deere', '5085GL',         'NO'),
    ('Agrícola',     'Tractor Utilitarios', 'Tractor Agrícola',           'John Deere', '5065E',          'NO'),
    ('Agrícola',     'Tractor Utilitarios', 'Tractor Agrícola',           'John Deere', '5075E',          'NO'),
    ('Agrícola',     'Tractor Utilitarios', 'Tractor Agrícola',           'John Deere', '5082E',          'NO'),
    ('Agrícola',     'Tractor Utilitarios', 'Tractor Agrícola',           'John Deere', '5076EF',         'SI-PER'),
    ('Agrícola',     'Tractor Especializados','Tractor Agrícola',         'John Deere', '3036EN',         'SI-PER'),
    ('Agrícola',     'Tractor Especializados','Tractor Agrícola',         'John Deere', '3036E',          'NO'),
    ('Agrícola',     'Tractor Mediano',     'Tractor Agrícola',           'John Deere', '6110E',          'NO'),
    ('Agrícola',     'Tractor Mediano',     'Tractor Agrícola',           'John Deere', '6105J',          'NO'),
    ('Agrícola',     'Tractor Mediano',     'Tractor Agrícola',           'John Deere', '6140J',          'NO'),
    ('Agrícola',     'Tractor Mediano',     'Tractor Agrícola',           'John Deere', '6155J',          'NO'),
    ('Agrícola',     'Tractor Mediano',     'Tractor Agrícola',           'John Deere', '6110D',          'SI-PER'),
    ('Agrícola',     'Tractor Mediano',     'Tractor Agrícola',           'John Deere', '6603',           'SI-PER'),
    ('Agrícola',     'Tractor Mediano',     'Tractor Agrícola',           'John Deere', '6115J',          'NO'),
    ('Agrícola',     'Tractor Mediano',     'Tractor Agrícola',           'John Deere', '6135J',          'NO'),
    ('Agrícola',     'Tractor Mediano',     'Tractor Agrícola',           'John Deere', '6150J',          'NO'),
    ('Agrícola',     'Tractor Mediano',     'Tractor Agrícola',           'John Deere', '6190J',          'NO'),
    ('Agrícola',     'Tractor Agricola',    'Tractor Agrícola',           'John Deere', '6403',           'SI-PER'),
    ('Agrícola',     'Tractor Grande',      'Tractor Agrícola',           'John Deere', '6803',           'NO'),
    ('Agrícola',     'Tractor Grande',      'Tractor Agrícola',           'John Deere', '7230J',          'SI-PER'),
    ('Agrícola',     'Tractor Grande',      'Tractor Agrícola Serie 6J',  'John Deere', 'Serie 6J',       'SI-PER'),
    ('Agrícola',     'Tractor Grande',      'Tractor Agrícola',           'John Deere', '6125E',          'SI-PER'),

    # ── IMPLEMENTOS ───────────────────────────────────────────────────────────
    ('Implementos',  'Aditamento',          'Martillo Hidráulico Excavadora',     'Hydrokhan', 'UG2100',      'SI-PER'),
    ('Implementos',  'Aditamento',          'Martillo Hidráulico Excavadora',     'Hydrokhan', 'UG3300',      'SI-PER'),
    ('Implementos',  'Aditamento',          'Martillo Hidráulico Retroexcavadora','Hydrokhan', 'UG400',       'SI-PER'),
    ('Implementos',  'Aditamento',          'Martillo Hidráulico Minicargador',   'Hydrokhan', 'UG350',       'SI-PER'),
    ('Implementos',  'Aditamento',          'Gancho de Excavadora',               'John Deere','350G-LC',     'NO'),
    ('Implementos',  'Aditamento',          'Ripper de Excavadora',               'John Deere','350G-LC',     'NO'),
    ('Implementos',  'Aditamento',          'Brazo Largo Excavadora',             'John Deere','210G',        'NO'),
    ('Implementos',  'Aditamento',          'Pulverizador / Atomizador',          'Gamma',     'Citrus Torre','SI-PER'),
    ('Implementos',  'Aditamento',          'Portabines',                         'Chaski',    'P 2500',      'NO'),
]

# ══════════════════════════════════════════════════════════════════════════════
# ESTILOS
# ══════════════════════════════════════════════════════════════════════════════
H_FILL  = PatternFill('solid', fgColor='1B4D35')
H_FONT  = Font(bold=True, color='FFFFFF', size=11)
N_FONT  = Font(size=10)
thin    = Side(style='thin', color='CCCCCC')
BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR     = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT     = Alignment(horizontal='left',   vertical='center', wrap_text=True)

SEC_FILL = {
    'Construcción': PatternFill('solid', fgColor='D6EAF8'),
    'Minería':      PatternFill('solid', fgColor='FDEDEC'),
    'Energía':      PatternFill('solid', fgColor='FEF9E7'),
    'Agrícola':     PatternFill('solid', fgColor='EAFAF1'),
    'Implementos':  PatternFill('solid', fgColor='F5EEF8'),
}

WEB_FILL = {
    'SI-AMBOS': PatternFill('solid', fgColor='A9DFBF'),  # verde
    'SI-PER':   PatternFill('solid', fgColor='AED6F1'),  # azul
    'SI-ARG':   PatternFill('solid', fgColor='FAD7A0'),  # naranja
    'NO':       PatternFill('solid', fgColor='FADBD8'),  # rojo
}
WEB_LABEL = {
    'SI-AMBOS': '✔ Perú + Argentina',
    'SI-PER':   '✔ Solo Perú',
    'SI-ARG':   '✔ Solo Argentina',
    'NO':       '✘ No está en web',
}

SEC_HDR = {
    'Construcción': PatternFill('solid', fgColor='1A5276'),
    'Minería':      PatternFill('solid', fgColor='7B241C'),
    'Energía':      PatternFill('solid', fgColor='7D6608'),
    'Agrícola':     PatternFill('solid', fgColor='1D6A39'),
    'Implementos':  PatternFill('solid', fgColor='6C3483'),
}

# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Brochure Desglosado'

headers = ['#', 'SECTOR', 'TIPO DE EQUIPO', 'EQUIPO / NOMBRE', 'MARCA', 'MODELO', 'EN PÁGINA WEB']
col_w   = [4,   14,       22,                30,                14,      18,        22            ]

for c, (h, w) in enumerate(zip(headers, col_w), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = H_FILL; cell.font = H_FONT
    cell.alignment = CTR; cell.border = BORDER
    ws.column_dimensions[get_column_letter(c)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = 'A2'

current_sector = None
row_num = 2
item_num = 1

for sector, tipo, equipo, marca, modelo, en_web in data:

    # ── Fila separadora de sector ────────────────────────────────────────────
    if sector != current_sector:
        current_sector = sector
        hdr_fill = SEC_HDR.get(sector, H_FILL)
        emoji = {'Construcción':'🏗️','Minería':'⛏️','Energía':'⚡','Agrícola':'🌾','Implementos':'🔧'}.get(sector,'')
        ws.merge_cells(f'A{row_num}:G{row_num}')
        cell = ws.cell(row=row_num, column=1, value=f'  {emoji}  {sector.upper()}')
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color='FFFFFF', size=12)
        cell.alignment = LFT
        cell.border = BORDER
        ws.row_dimensions[row_num].height = 20
        row_num += 1
        item_num = 1   # reset counter per sector

    sf   = SEC_FILL.get(sector, PatternFill())
    wf   = WEB_FILL.get(en_web, PatternFill())
    wlbl = WEB_LABEL.get(en_web, '')

    vals = [item_num, sector, tipo, equipo, marca, modelo, wlbl]

    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row_num, column=c, value=v)
        cell.border = BORDER
        cell.font   = N_FONT
        cell.fill   = wf if c == 7 else sf
        cell.alignment = CTR if c in (1,) else LFT

    ws.row_dimensions[row_num].height = 17
    row_num  += 1
    item_num += 1

# Autofilter (sin la fila 1 de header, porque hay filas de sector intercaladas)
ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

# ── Hoja resumen ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Resumen')
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 14

# Conteos
from collections import Counter
conteos = {}
for sector, tipo, equipo, marca, modelo, en_web in data:
    if sector not in conteos:
        conteos[sector] = {'SI-AMBOS':0,'SI-PER':0,'SI-ARG':0,'NO':0,'total':0}
    conteos[sector][en_web] += 1
    conteos[sector]['total'] += 1

# Header resumen
for c, h in enumerate(['SECTOR','EN WEB (Ambos)','Solo Perú','No en web','TOTAL'], 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.fill = H_FILL; cell.font = H_FONT; cell.alignment = CTR; cell.border = BORDER

total_all = 0
for r, (sector, cnt) in enumerate(conteos.items(), 2):
    hf = SEC_HDR.get(sector, H_FILL)
    vals2 = [sector, cnt['SI-AMBOS'], cnt['SI-PER'], cnt['NO'], cnt['total']]
    for c, v in enumerate(vals2, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.font   = N_FONT
        cell.fill   = PatternFill('solid', fgColor=hf.fgColor) if c==1 else PatternFill()
        if c == 1: cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = CTR if c > 1 else LFT
    total_all += cnt['total']
    ws2.row_dimensions[r].height = 18

# Fila total
tr = len(conteos) + 2
ws2.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True, size=11)
ws2.cell(row=tr, column=5, value=total_all).font = Font(bold=True, size=11)
for c in range(1,6):
    ws2.cell(row=tr, column=c).border = BORDER
    ws2.cell(row=tr, column=c).fill = PatternFill('solid', fgColor='D5E8D4')
    ws2.cell(row=tr, column=c).alignment = CTR

wb.save('brochure_desglosado.xlsx')
total = len(data)
en_web = sum(1 for d in data if d[5] != 'NO')
no_web = sum(1 for d in data if d[5] == 'NO')
print(f'Excel generado OK')
print(f'Total modelos: {total}')
print(f'En web: {en_web}  |  No en web: {no_web}')
