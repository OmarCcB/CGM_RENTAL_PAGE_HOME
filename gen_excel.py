import sqlite3, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# ── Paleta ───────────────────────────────────────────────────────────────────
H_FILL   = PatternFill('solid', fgColor='1B4D35')
INACT_F  = PatternFill('solid', fgColor='F2F3F4')
PROP_F   = PatternFill('solid', fgColor='FFF8E1')   # amarillo pálido → propuesta
PERU_F   = PatternFill('solid', fgColor='D6E4F7')   # azul suave  → Perú
ARG_F    = PatternFill('solid', fgColor='FFF3CD')   # amarillo    → Argentina
SEC_COLORS = {
    'Construcción':          ('D6EAF8', 'C0D8EE'),   # (Peru, Arg)
    'Agrícola':              ('EAFAF1', 'D4F0E2'),
    'Energía':               ('FEF9E7', 'FDF0C8'),
    'Mediana Minería':       ('FDEDEC', 'FBDAD8'),
    'Construcción, Energía': ('EBF5FB', 'D5EAF5'),
}

def sec_fill(unidad, country):
    pair = SEC_COLORS.get(unidad, ('F9F9F9', 'F0F0F0'))
    return PatternFill('solid', fgColor=pair[0] if country == 'per' else pair[1])

H_FONT = Font(bold=True, color='FFFFFF', size=11)
N_FONT = Font(size=10)
thin   = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR    = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT    = Alignment(horizontal='left',   vertical='center', wrap_text=True)

# ── URL builder ──────────────────────────────────────────────────────────────
TIPO_SLUG = {
    'Excavadora':            'excavadora',
    'Cargador Frontal':      'cargador-frontal',
    'Tractor de Orugas':     'tractor-de-orugas',
    'Rodillo Compactador':   'rodillo-compactador',
    'Rodillo Tándem':        'rodillo-compactador',
    'Motoniveladora':        'motoniveladora',
    'Retroexcavadora':       'retroexcavadora',
    'Minicargador':          'minicargador',
    'Camion Cisterna':       'camion-cisterna',
    'Camión Cisterna':       'camion-cisterna',
    'Camion Grua':           'camion-grua',
    'Camión Grúa':           'camion-grua',
    'Camión Volquete':       'camion-volquete',
    'Torre de Iluminacion':  'torre-de-iluminacion',
    'Torre de Iluminación':  'torre-de-iluminacion',
    'Aditamento':            'aditamentos',
    'Micropavimentadora':    'micropavimentadora',
    'Pavimentadora':         'pavimentadora',
    'Autohormigonera':       'autohormigonera',
    'Chancadora':            'chancadora',
    'Zaranda':               'zaranda',
    'Faja Transportadora':   'faja-transportadora',
    'Compresora':            'compresora',
    'Generador':             None,   # va directo a /energia/
}
UNIDAD_SLUG = {
    'Construcción':              'construccion',
    'Agrícola':                  'agricola',
    'Energía':                   'energia',
    'Mediana Minería':           'mineria',
    'Construcción, Energía':     'construccion',
}

def build_url(country, tags, unidad, tipo, slug):
    u  = UNIDAD_SLUG.get(unidad, 'construccion')
    if tags == 'usados':
        return f'/{country}/usados/{slug}/'
    ts = TIPO_SLUG.get(tipo)
    if ts and u == 'construccion':
        return f'/{country}/alquiler/construccion/{ts}/{slug}/'
    return f'/{country}/alquiler/{u}/{slug}/'

# ── Tipos por sector (para dropdown dependiente) ─────────────────────────────
TIPOS_POR_SECTOR = {
    'Construcción': [
        'Excavadora', 'Cargador Frontal', 'Retroexcavadora', 'Minicargador',
        'Motoniveladora', 'Tractor de Orugas', 'Rodillo Compactador', 'Rodillo Tándem',
        'Camion Cisterna', 'Camión Volquete', 'Camion Grua', 'Torre de Iluminacion',
        'Micropavimentadora', 'Pavimentadora', 'Autohormigonera',
        'Chancadora', 'Zaranda', 'Faja Transportadora', 'Aditamento',
    ],
    'Agrícola': [
        'Tractor Agricola', 'Tractor Grande', 'Tractor Mediano',
        'Tractor Utilitarios', 'Tractor Especializados', 'Aditamento',
    ],
    'Energía': [
        'Generador', 'Torre de Iluminación', 'Compresora',
    ],
    'Mediana Minería': [
        'Excavadora', 'Aditamento',
    ],
}

import re as _re

def to_slug(text):
    t = text.lower()
    t = t.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ñ','n')
    t = _re.sub(r'[^a-z0-9]+', '-', t).strip('-')
    return t

# ── DB ───────────────────────────────────────────────────────────────────────
conn = sqlite3.connect('cgm.db')
cur  = conn.cursor()
cur.execute(
    'SELECT id,slug,nombre,marca,tipo,unidad,tags,activo,show_arg '
    'FROM products ORDER BY unidad,tipo,nombre'
)
db_rows = cur.fetchall()
conn.close()

# ── Productos propuestos (no están en la web, solo Perú, solo Alquiler) ──────
# Formato: (nombre, marca, tipo, unidad, slug_sugerido)
propuestos = [
    # CONSTRUCCIÓN
    ('CARGADOR FRONTAL 844L',              'JOHN DEERE', 'Cargador Frontal',    'Construcción', 'cargador-frontal-844l'),
    ('RODILLO COMPACTADOR HC 200',         'HAMM',       'Rodillo Compactador', 'Construcción', 'rodillo-compactador-hamm-hc200'),
    ('CAMIÓN VOLQUETE CF 480',             'DAF',        'Camión Volquete',      'Construcción', 'camion-volquete-daf-cf480'),
    ('MICROPAVIMENTADORA CF 85.460',       'Bergkamp',   'Micropavimentadora',   'Construcción', 'micropavimentadora-bergkamp-cf85460'),
    ('PAVIMENTADORA SUPER 1800-3i',        'Vogele',     'Pavimentadora',        'Construcción', 'pavimentadora-vogele-super-1800-3i'),
    ('PAVIMENTADORA 1400',                 'Vogele',     'Pavimentadora',        'Construcción', 'pavimentadora-vogele-1400'),
    ('AUTOHORMIGONERA DB 460',             'Fiori',      'Autohormigonera',      'Construcción', 'autohormigonera-fiori-db460'),
    ('CHANCADORA PRIMARIA MC 110 EVO2',    'Kleemann',   'Chancadora',           'Construcción', 'chancadora-primaria-kleemann-mc110-evo2'),
    ('CHANCADORA SECUNDARIA MC 90 EVO2',   'Kleemann',   'Chancadora',           'Construcción', 'chancadora-secundaria-kleemann-mc90-evo2'),
    ('ZARANDA',                            'Kleemann',   'Zaranda',              'Construcción', 'zaranda-kleemann'),
    ('FAJA TRANSPORTADORA MBT 24',         'Kleemann',   'Faja Transportadora',  'Construcción', 'faja-transportadora-kleemann-mbt24'),
    # ENERGÍA
    ('GENERADOR AKSA AK70CUST',            'AKSA',       'Generador',            'Energía',      'generador-aksa-ak70cust'),
    ('GENERADOR AKSA AK90CUST',            'AKSA',       'Generador',            'Energía',      'generador-aksa-ak90cust'),
    ('GENERADOR AKSA AK320CUST',           'AKSA',       'Generador',            'Energía',      'generador-aksa-ak320cust'),
    ('GENERADOR TEKSAN TK55CUST',          'TEKSAN',     'Generador',            'Energía',      'generador-teksan-tk55cust'),
    ('GENERADOR TEKSAN TK180PELS',         'TEKSAN',     'Generador',            'Energía',      'generador-teksan-tk180pels'),
    ('GENERADOR TEKSAN TK220PELS',         'TEKSAN',     'Generador',            'Energía',      'generador-teksan-tk220pels'),
    ('GENERADOR TEKSAN TK280PELS',         'TEKSAN',     'Generador',            'Energía',      'generador-teksan-tk280pels'),
    ('GENERADOR TEKSAN TK500PELS',         'TEKSAN',     'Generador',            'Energía',      'generador-teksan-tk500pels'),
    ('GENERADOR TEKSAN TK725CUST',         'TEKSAN',     'Generador',            'Energía',      'generador-teksan-tk725cust'),
    ('TORRE DE ILUMINACIÓN GLT4-M',        'Generac Vtevo','Torre de Iluminación','Energía',     'torre-de-iluminacion-generac-glt4m'),
    ('TORRE DE ILUMINACIÓN ATLAS COPCO HILIGHT V5+', 'Atlas Copco','Torre de Iluminación','Energía', 'torre-de-iluminacion-atlas-copco-hilight-v5'),
    ('COMPRESORA XAS186',                  'Atlas Copco','Compresora',           'Energía',      'compresora-atlas-copco-xas186'),
    # AGRÍCOLA
    ('TRACTOR AGRÍCOLA 7M-230',            'JOHN DEERE', 'Tractor Grande',       'Agrícola',     'tractor-agricola-7m-230'),
    ('TRACTOR AGRÍCOLA 5090EH',            'JOHN DEERE', 'Tractor Utilitarios',  'Agrícola',     'tractor-agricola-5090eh'),
    ('TRACTOR AGRÍCOLA 5080EN',            'JOHN DEERE', 'Tractor Utilitarios',  'Agrícola',     'tractor-agricola-5080en'),
    ('TRACTOR AGRÍCOLA 5085GL',            'JOHN DEERE', 'Tractor Utilitarios',  'Agrícola',     'tractor-agricola-5085gl'),
    ('TRACTOR AGRÍCOLA 5065E',             'JOHN DEERE', 'Tractor Utilitarios',  'Agrícola',     'tractor-agricola-5065e'),
    ('TRACTOR AGRÍCOLA 5075E',             'JOHN DEERE', 'Tractor Utilitarios',  'Agrícola',     'tractor-agricola-5075e'),
    ('TRACTOR AGRÍCOLA 5082E',             'JOHN DEERE', 'Tractor Utilitarios',  'Agrícola',     'tractor-agricola-5082e'),
    ('TRACTOR AGRÍCOLA 3036E',             'JOHN DEERE', 'Tractor Especializados','Agrícola',    'tractor-agricola-3036e'),
    ('TRACTOR AGRÍCOLA 6110E',             'JOHN DEERE', 'Tractor Mediano',      'Agrícola',     'tractor-agricola-6110e'),
    ('TRACTOR AGRÍCOLA 6105J',             'JOHN DEERE', 'Tractor Mediano',      'Agrícola',     'tractor-agricola-6105j'),
    ('TRACTOR AGRÍCOLA 6140J',             'JOHN DEERE', 'Tractor Grande',       'Agrícola',     'tractor-agricola-6140j'),
    ('TRACTOR AGRÍCOLA 6155J',             'JOHN DEERE', 'Tractor Grande',       'Agrícola',     'tractor-agricola-6155j'),
    ('TRACTOR AGRÍCOLA 6115J',             'JOHN DEERE', 'Tractor Mediano',      'Agrícola',     'tractor-agricola-6115j'),
    ('TRACTOR AGRÍCOLA 6135J',             'JOHN DEERE', 'Tractor Grande',       'Agrícola',     'tractor-agricola-6135j'),
    ('TRACTOR AGRÍCOLA 6150J',             'JOHN DEERE', 'Tractor Grande',       'Agrícola',     'tractor-agricola-6150j'),
    ('TRACTOR AGRÍCOLA 6190J',             'JOHN DEERE', 'Tractor Grande',       'Agrícola',     'tractor-agricola-6190j'),
    ('TRACTOR AGRÍCOLA 6803',              'JOHN DEERE', 'Tractor Grande',       'Agrícola',     'tractor-agricola-6803'),
    # IMPLEMENTOS (Construcción)
    ('GANCHO DE EXCAVADORA 350G-LC',       'JOHN DEERE', 'Aditamento',           'Construcción', 'gancho-excavadora-350g-lc'),
    ('RIPPER DE EXCAVADORA 350G-LC',       'JOHN DEERE', 'Aditamento',           'Construcción', 'ripper-excavadora-350g-lc'),
    ('BRAZO LARGO EXCAVADORA 210G',        'JOHN DEERE', 'Aditamento',           'Construcción', 'brazo-largo-excavadora-210g'),
    # IMPLEMENTOS (Agrícola)
    ('PORTABINES P 2500',                  'Chaski',     'Aditamento',           'Agrícola',     'portabines-chaski-p2500'),
]

# ── Expandir filas DB: duplicar los que están en ambos países ────────────────
expanded = []
for row in db_rows:
    pid, slug, nombre, marca, tipo, unidad, tags, activo, show_arg = row
    peru_ok = activo == 1
    arg_ok  = activo == 1 and show_arg == 1

    if peru_ok and arg_ok:
        expanded.append(('per', 'Perú',        row, 'activo'))
        expanded.append(('arg', 'Argentina',   row, 'activo'))
    elif peru_ok:
        expanded.append(('per', 'Perú',        row, 'activo'))
    elif arg_ok:
        expanded.append(('arg', 'Argentina',   row, 'activo'))
    else:
        expanded.append(('ina', '— Inactivo —', row, 'inactivo'))

# ── Agregar propuestos (solo Perú, solo Alquiler) ────────────────────────────
for nombre, marca, tipo, unidad, slug in propuestos:
    fake_row = (None, slug, nombre, marca, tipo, unidad, 'alquiler', 1, 0)
    expanded.append(('per', 'Perú', fake_row, 'propuesta'))

# ════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── Hoja _Listas (oculta) ────────────────────────────────────────────────────
wl = wb.create_sheet('_Listas')
wl.sheet_state = 'hidden'

# Columna A: Países
for i, v in enumerate(['Perú', 'Argentina'], 1):
    wl.cell(row=i, column=1, value=v)

# Columna B: Modalidad_Peru
for i, v in enumerate(['Alquiler', 'Usados'], 1):
    wl.cell(row=i, column=2, value=v)

# Columna C: Modalidad_Argentina
wl.cell(row=1, column=3, value='Alquiler')

# Columna D: Sector_Peru_Alquiler
for i, v in enumerate(['Construcción', 'Agrícola', 'Energía', 'Mediana Minería'], 1):
    wl.cell(row=i, column=4, value=v)

# Columna E: Sector_Peru_Usados
for i, v in enumerate(['Construcción', 'Agrícola'], 1):
    wl.cell(row=i, column=5, value=v)

# Columna F: Sector_Arg_Alquiler
wl.cell(row=1, column=6, value='Construcción')

# Tipos por sector — columnas G..J
tipo_cols = {
    'Construcción':   7,   # G
    'Agrícola':       8,   # H
    'Energía':        9,   # I
    'Mediana Minería':10,  # J
}
tipo_col_letter = {s: get_column_letter(c) for s, c in tipo_cols.items()}
tipo_rows = {}
for sector, col in tipo_cols.items():
    tipos = TIPOS_POR_SECTOR[sector]
    for i, v in enumerate(tipos, 1):
        wl.cell(row=i, column=col, value=v)
    tipo_rows[sector] = len(tipos)

# Named ranges
sheet_ref = "'_Listas'"
dn_map = {
    'Paises':                f"{sheet_ref}!$A$1:$A$2",
    'Modalidad_Peru':        f"{sheet_ref}!$B$1:$B$2",
    'Modalidad_Argentina':   f"{sheet_ref}!$C$1:$C$1",
    'Sector_Peru_Alquiler':  f"{sheet_ref}!$D$1:$D$4",
    'Sector_Peru_Usados':    f"{sheet_ref}!$E$1:$E$2",
    'Sector_Arg_Alquiler':   f"{sheet_ref}!$F$1:$F$1",
    # Tipos
    'Tipos_Construccion':    f"{sheet_ref}!$G$1:$G${tipo_rows['Construcción']}",
    'Tipos_Agricola':        f"{sheet_ref}!$H$1:$H${tipo_rows['Agrícola']}",
    'Tipos_Energia':         f"{sheet_ref}!$I$1:$I${tipo_rows['Energía']}",
    'Tipos_Mineria':         f"{sheet_ref}!$J$1:$J${tipo_rows['Mediana Minería']}",
}
for name, ref in dn_map.items():
    wb.defined_names[name] = DefinedName(name, attr_text=ref)

# ── Hoja principal ───────────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Productos CGM Rental'

# Columnas: ID | NOMBRE | MARCA | PAÍS | MODALIDAD | SECTOR | TIPO DE EQUIPO | ESTADO | URL
headers = ['ID', 'NOMBRE', 'MARCA', 'PAÍS', 'MODALIDAD', 'SECTOR', 'TIPO DE EQUIPO', 'ESTADO', 'URL']
col_w   = [5,    42,       16,      12,      13,           22,        22,               10,       55  ]
# D=4=País  E=5=Modalidad  F=6=Sector  G=7=Tipo

for c, (h, w) in enumerate(zip(headers, col_w), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = H_FILL; cell.font = H_FONT
    cell.alignment = CTR; cell.border = BORDER
    ws.column_dimensions[get_column_letter(c)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = 'A2'

# ── Llenar filas ─────────────────────────────────────────────────────────────
PROP_FONT = Font(size=10, italic=True, color='7D6608')

for r, (country_code, pais_label, row, estado_flag) in enumerate(expanded, 2):
    pid, slug, nombre, marca, tipo, unidad, tags, activo, show_arg = row

    url = build_url(country_code, tags, unidad, tipo, slug) if country_code != 'ina' else '—'

    if estado_flag == 'propuesta':
        estado_lbl = '⭐ PROPUESTA'
        row_fill   = PROP_F
        row_font   = PROP_FONT
    elif estado_flag == 'inactivo':
        estado_lbl = 'Inactivo'
        row_fill   = INACT_F
        row_font   = N_FONT
    else:
        estado_lbl = 'Activo'
        row_fill   = sec_fill(unidad, country_code)
        row_font   = N_FONT

    mod = tags.capitalize()
    vals = [pid if pid else '—', nombre, marca, pais_label, mod, unidad, tipo, estado_lbl, url]

    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.font   = row_font
        cell.fill   = row_fill
        cell.alignment = CTR if c in (1, 4, 5, 8) else LFT

    ws.row_dimensions[r].height = 18

nrows = len(expanded) + 1

# ── Data Validations ─────────────────────────────────────────────────────────
# D = País
dv_pais = DataValidation(
    type='list', formula1='Paises', showDropDown=False,
    showErrorMessage=True,
    error='Solo se permite: Perú o Argentina', errorTitle='País inválido')
dv_pais.sqref = f'D2:D{nrows}'
ws.add_data_validation(dv_pais)

# E = Modalidad — depende de País
dv_mod = DataValidation(
    type='list',
    formula1='IF(D2="Argentina",Modalidad_Argentina,Modalidad_Peru)',
    showDropDown=False, showErrorMessage=True,
    error='Argentina solo admite Alquiler.', errorTitle='Modalidad inválida')
dv_mod.sqref = f'E2:E{nrows}'
ws.add_data_validation(dv_mod)

# F = Sector — depende de País + Modalidad
dv_sec = DataValidation(
    type='list',
    formula1='IF(D2="Argentina",Sector_Arg_Alquiler,IF(E2="Usados",Sector_Peru_Usados,Sector_Peru_Alquiler))',
    showDropDown=False, showErrorMessage=True,
    error='Sector no permitido para esta combinación.', errorTitle='Sector inválido')
dv_sec.sqref = f'F2:F{nrows}'
ws.add_data_validation(dv_sec)

# G = Tipo de Equipo — depende de Sector (F)
dv_tipo = DataValidation(
    type='list',
    formula1='IF(F2="Construcción",Tipos_Construccion,IF(F2="Agrícola",Tipos_Agricola,IF(F2="Energía",Tipos_Energia,Tipos_Mineria)))',
    showDropDown=False, showErrorMessage=True,
    error='Tipo de equipo no válido para el sector seleccionado.', errorTitle='Tipo inválido')
dv_tipo.sqref = f'G2:G{nrows}'
ws.add_data_validation(dv_tipo)

# Autofilter
ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

# ── Hoja Referencia ──────────────────────────────────────────────────────────
wr = wb.create_sheet('Referencia')
LFT2 = Alignment(horizontal='left', vertical='center', wrap_text=True)
RED_F = PatternFill('solid', fgColor='FADBD8')
GRN_F = PatternFill('solid', fgColor='D5E8D4')

ref_rows = [
    ('REGLAS DE COMBINACIONES VÁLIDAS', '', '', '', 'title'),
    ('', '', '', '', None),
    ('PAÍS', 'MODALIDAD', 'SECTOR', 'TIPOS DE EQUIPO DISPONIBLES', 'header'),
    ('Perú', 'Alquiler', 'Construcción',
     'Excavadora | Cargador Frontal | Retroexcavadora | Minicargador | Motoniveladora | Tractor de Orugas | Rodillo Compactador | Camion Cisterna | Camion Grua | Torre de Iluminacion | Aditamento', 'ok'),
    ('Perú', 'Alquiler', 'Agrícola',
     'Tractor Agricola | Tractor Grande | Tractor Mediano | Tractor Utilitarios | Tractor Especializados | Aditamento', 'ok'),
    ('Perú', 'Alquiler', 'Energía',
     'Generador | Torre de Iluminacion', 'ok'),
    ('Perú', 'Alquiler', 'Mediana Minería',
     'Excavadora | Aditamento', 'ok'),
    ('Perú', 'Usados', 'Construcción',
     'Excavadora | Cargador Frontal | Retroexcavadora | Minicargador | Motoniveladora | Tractor de Orugas | Rodillo Compactador | Camion Cisterna | Camion Grua | Torre de Iluminacion | Aditamento', 'ok'),
    ('Perú', 'Usados', 'Agrícola',
     'Tractor Agricola | Tractor Grande | Tractor Mediano | Tractor Utilitarios | Tractor Especializados | Aditamento', 'ok'),
    ('Argentina', 'Alquiler', 'Construcción',
     'Excavadora | Cargador Frontal | Retroexcavadora | Minicargador | Motoniveladora | Tractor de Orugas | Rodillo Compactador | Camion Cisterna | Camion Grua | Torre de Iluminacion | Aditamento', 'ok'),
    ('Argentina', 'Usados', '— NO APLICA —', '— NO APLICA —', 'no'),
]

for i, row_data in enumerate(ref_rows, 1):
    *vals, style = row_data
    for c, v in enumerate(vals, 1):
        cell = wr.cell(row=i, column=c, value=v)
        cell.alignment = LFT2
        if style == 'title':
            cell.font = Font(bold=True, color='1B4D35', size=13)
        elif style == 'header':
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill('solid', fgColor='1B4D35')
        elif style == 'ok':
            cell.font = Font(size=10)
            cell.fill = GRN_F
        elif style == 'no':
            cell.font = Font(size=10)
            cell.fill = RED_F
        else:
            cell.font = Font(size=10)
        cell.border = BORDER

wr.column_dimensions['A'].width = 12
wr.column_dimensions['B'].width = 12
wr.column_dimensions['C'].width = 18
wr.column_dimensions['D'].width = 90

wb.save('productos_cgm_rental.xlsx')
print(f'Excel generado OK — {len(expanded)} filas ({len(db_rows)} productos, algunos duplicados por país)')
